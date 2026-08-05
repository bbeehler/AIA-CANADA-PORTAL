import pandas as pd
from openpyxl import load_workbook

from aia_portal.exports import excel_report_bytes, pdf_report_bytes


def test_excel_export_contains_details_and_data():
    payload = excel_report_bytes(pd.DataFrame({"region": ["Ontario"], "value": [4.1]}), title="Test report")
    workbook = load_workbook(filename=__import__("io").BytesIO(payload), read_only=True)
    assert workbook.sheetnames == ["Report details", "Data"]
    assert workbook["Data"]["A2"].value == "Ontario"


def test_pdf_export_is_a_pdf():
    payload = pdf_report_bytes(pd.DataFrame({"region": ["Ontario"], "value": [4.1]}), title="Test report")
    assert payload.startswith(b"%PDF")
    assert len(payload) > 1000
